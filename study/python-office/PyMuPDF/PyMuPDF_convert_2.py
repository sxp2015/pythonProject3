#!/usr/bin/python3
# coding:utf-8            
#
# Copyright (C) 2023 - 2023 Sunny， Inc. All Rights Reserved 
#
# @Time    : 2023-05-26 21:12
# @Author  : Code_By_Sunny_Sun
# @Email   : 240746804@qq.com
# @File    : conver_tesst.py
# @IDE     : PyCharm
# 导入必要的库
import os  # 处理文件和目录路径相关的功能
import fitz  # 操作 PDF 文件的库 PyMuPDF
import io  # 读取二进制数据的库
from PIL import Image  # 处理图片的库
import tabula  # 读取表格数据的库
import pandas as pd  # 处理表格数据的库
from datetime import datetime
from openpyxl import Workbook

# PDF 文件路径
pdf_file_path = os.path.join('pdf_files/查验报告.pdf')
doc = fitz.open(pdf_file_path)


# 获取PDF中所有页面的图片数量
def get_images_count():
    # 先找到起始页row==‘1’，作为开始点
    start_page = None

    # 从开始点统计到结束点row==‘1’为止，累加所有的图片数量
    end_page = None
    images_count = 0

    for page_num, page in enumerate(doc):
        rows = page.get_text("table").split("\n")
        print('rows:', rows)
        if '1' in rows:
            start_page = page_num
            # print('start_page:', start_page)

            # 遍历每一行
            for i, row in enumerate(rows):
                # 如果单元格图片
                cells = row.strip().split("\t")
                if cells:
                    if '1' in cells:
                        end_page = i
                        print('标记点索引:', i)
                        # 获取第一行的每个单元格
                        for j, cell in enumerate(cells[:end_page]):
                            # 使用 search_for 方法查找 PDF 中的单元格
                            text_result = page.search_for(cell)
                            image_result = page.get_images(cell)
                            print('image_result', image_result)
                            print('image_result:', len(image_result))
                    #  print('text_result:', text_result)

        # 如果当前页数超过总页数，则跳出循环
        if page_num >= len(doc):
            break

    if start_page is None:
        raise Exception('未找到起始页')

    # 从开始点统计到结束点row==‘1’为止，累加所有的图片数量
    # end_page = None
    # images_count = 0
    # for page_num, page in enumerate(doc[start_page:], start=start_page):
    #     rows = page.get_text("table")
    #     print('rows:', rows)
    #     if '1' in rows:
    #         end_page = page_num
    #         break
    #     image_list = page.get_images()
    #     if image_list:
    #         for img in image_list:
    #             xref = img[0]
    #             pix = fitz.Pixmap(doc, xref)
    #             if pix.w >= 1500 and pix.h >= 1100:
    #                 images_count += 1
    #             pix = None
    # if end_page is None:
    #     raise Exception('未找到结束页')
    #
    # return images_count


# 提取表格单元格中信息的方法
def extract_table_cells(page):
    # 调用 get_text("table") 方法获取表格数据
    table_data = page.get_text("table")
    # 输出表格数据
    print('【table_data】: ', table_data)

    # 如果表格数据为空，则返回空列表
    if not table_data:
        return []

    # 将表格数据按行分割
    rows = table_data.split("\n")

    # 用于存储单元格数据和位置信息的列表
    table_cells = []
    start_page = None
    # 遍历表格数据的每一行
    for i, row in enumerate(rows):

        # 提取房号
        if '房　　号' in row:
            row_cell_value, index = row, i
            house_number = rows[index + 1]
            table_cells.append({
                "house_number": house_number
            })
            print(f'house_number房号是:{house_number}, index下标索引是:{index + 1}')

        # 获取列将每一行的单元格，按制表符（\t）分割成多个单元格

        cells = row.strip().split("\t")

        # 如果该行不为空，则遍历该行中的所有单元格
        if cells:

            # 获取第一行的每个单元格
            for j, cell in enumerate(cells):
                # 使用 search_for 方法查找 PDF 中的单元格
                text_result = page.search_for(cell)
                image_result = page.get_images(cell)

                # 如果单元格有文本的处理
                if text_result and cell:
                    # print(f'第{page.number + 1}页,行索引号：{i + 1}，单元格文本内容是：{cell}')
                    table_cell = {"value": cell, "page_num": page.number + 1, "row:": i + 1,
                                  "column:": j + 1}
                    table_cells.append(table_cell)

                # 如果单元格有图片的处理
                if image_result and row == '1':
                    # print('image_result', image_result)

                    # 获取检查部位
                    check_part = rows[i - 1]

                    print(f'==检查部位是==:{check_part}')
                    # 定义一个变量统计满足条件的图片有多少
                    img_count = 0

                    for m, img in enumerate(image_result):
                        # 将对象转换为 Rect 类型，并获取其位置信息
                        # search_for 方法返回的结果是一个列表，其中每个元素都是一个包含四个元素的元组，
                        # 分别表示匹配项的左、上、右、下四个角的坐标。这四个坐标可以用来创建一个 fitz.Rect 对象，
                        # 从而表示该匹配项在 PDF 页面上的位置和大小。因此，代码中的 bbox = fitz.Rect(obj[:4])
                        # 表示将第一个匹配项的四个坐标传递给 fitz.Rect 构造函数，创建一个 Rect 对象并将其赋值给 bbox 变量。
                        # 结果就是 bbox 变量包含了这个匹配项在页面上的位置和大小信息，可以用于后续的处理。
                        bbox = fitz.Rect(img[:4])
                        # 记录当前图片的索引,根据图片尺寸和大小判断是不是需要的图片

                        if bbox.x1 >= 1500 and bbox.y1 >= 1100:
                            img_count += 1
                            print(f'==图片索引=:{m + 1},图片数量是:{img_count}')
                        # 将单元格的数据和位置信息添加到列表中
                        table_cells.append({
                            "value": f"图片{m + 1}-{img}",
                            "page_num": page.number + 1,
                            "check_part_name": check_part,
                            "left": bbox.x0,
                            "top": bbox.y0,
                            "right": bbox.x1,
                            "bottom": bbox.y1
                        })

    return table_cells

    # 遍历表格数据的每一列


# 保存PDF中的每页所有图片

def save_images_in_pdf(pdf_path):
    # 创建保存图片的文件夹
    if not os.path.exists("pdf_image"):
        os.mkdir("pdf_image")
    doc = fitz.open(pdf_path)

    # 遍历PDF文档每一页
    for page_num, page in enumerate(doc):

        # 保存图片列表
        image_list = page.get_images()
        if image_list:
            print(f"Page {page_num + 1} contains {len(image_list)} image(s):")
            for image_index, image in enumerate(image_list):
                xref = image[0]
                pix = fitz.Pixmap(doc, xref)
                if pix.w >= 1500 and pix.h >= 1100:
                    image_dir = "pdf_image"
                    if not os.path.exists(image_dir):
                        os.makedirs(image_dir)
                    image_path = os.path.join(
                        image_dir, f"查验报告_page{page_num + 1}_image{image_index + 1}.png"
                    )
                    # pix.save(image_path)
                    print(f"\t{image_path}")
                pix = None
        else:
            print(f"Page {page_num + 1} does not contain any images.")

    doc.close()


# 创建保存数据到Excel的方法
def create_extract_excel():
    # 创建工作簿
    wb = Workbook()

    # 获取当前时间并格式化
    dt = datetime.now().strftime("%Y%m%d_%H%M%S")

    # 确定文件夹路径和文件名
    folder_path = "./generated_excel"
    filename = f"{folder_path}/extract_excel_{dt}.xlsx"

    # 用于存储所有表格单元格数据的列表
    table_cells_list = []

    # 如果文件夹不存在，则创建
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    # 获取默认的第一个工作表对象
    ws = wb.active

    # 写入表头数据
    headers = ["序号", "创建日期", "楼栋", "房号", "部位", "检查项", "问题描述", "问题照片", "检查人", "责任单位", "工种", "工单状态", "销项照片", "最后修改人",
               "最后修改时间", "备注"]

    # 遍历PDF文档每一页
    for page_num, page in enumerate(doc):
        # 对当前页执行表格提取
        table_cells_list = extract_table_cells(page)

        # 如果存在表格，则打印表格数据
        if table_cells_list:
            print('table_cells_list', table_cells_list)
        else:
            print(f"No table found on page {page.number}.")

    for i, header in enumerate(headers):
        ws.cell(1, i + 1, header)

    for j, table_cell in enumerate(table_cells_list):
        ws.cell(2, j + 1, table_cell["value"])

    # 保存工作簿
    wb.save(filename)

    print(f"成功创建工作簿 {filename}！")


if __name__ == "__main__":
    # save_images_in_pdf(pdf_file_path)
    print('*' * 80)
    # create_extract_excel()
    image_count = get_images_count()
    print('image_count', image_count)
    # extract_table_images_and_text(pdf_file_path)
