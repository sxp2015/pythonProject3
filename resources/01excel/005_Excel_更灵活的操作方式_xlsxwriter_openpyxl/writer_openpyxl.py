# -*- coding: utf-8 -*-
# ⭐Python 60套学习资源：http://t.cn/A6xASARf
# 📱公众号 :程序员晚枫 读者交流群：http://www.python4office.cn/wechat-group/
# 🏠2022最新视频：1行代码，实现自动化办公：https://www.bilibili.com/video/BV1pT4y1k7FH


# import xlwt
#
# workbook = xlwt.Workbook()
# sheet0 = workbook.add_sheet('sheet0')
# for i in range(0,300):
#     sheet0.write(0,i,i)
# workbook.save('num.xls')

# 不带格式
import xlsxwriter as xw
workbook = xw.Workbook('number.xlsx')
sheet0 = workbook.add_worksheet('sheet0')
for i in range(0,300):
    sheet0.write(0,i,i)
workbook.close()


# 性能不稳定
import openpyxl
workbook = openpyxl.load_workbook('number.xlsx')
sheet0 = workbook['sheet0']
sheet0['B3']= '2'
sheet0['C2']= '4'
sheet0['D7']= '3'
workbook.save('num_open.xlsx')

